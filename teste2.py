import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

print(os.getcwd())

abas = ["Procedimentos", "Especialidades", "Cadastro", "Cálculos"]

unidas = {}


pasta_arquivos = 'C:/Users/Arthur Coutinho/Desktop/Arthur Coutinho/teste/'

arquivos = [f for f in os.listdir(pasta_arquivos) if f.endswith('.xlsx')]

for aba in abas:
    dfs = []
    for arquivo in arquivos:
        df = pd.read_excel(os.path.join(pasta_arquivos, arquivo), sheet_name=aba)
        dfs.append(df)

    unidas[aba] = pd.concat(dfs, ignore_index=True)

wb = Workbook()

for aba, df in unidas.items():
    ws = wb.create_sheet(title=aba)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

wb.save('C:/Users/Arthur Coutinho/Desktop/final/final.xlsx')

print("Planilhas unidas com sucesso!")
