import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

print(os.getcwd())

abas = ["Procedimentos", "Especialidades", "Cadastro", "Cálculos"]

unidas = {}

for aba in abas:
    dfs = []
    for i in range(1, 4):
        df = pd.read_excel(f'C:/Users/Arthur Coutinho/Desktop/Arthur Coutinho/teste/teste{i}.xlsx', sheet_name=aba)
        dfs.append(df)

    unidas[aba] = pd.concat(dfs, ignore_index=True)

wb = Workbook()

for aba, df in unidas.items():
    ws = wb.create_sheet(title=aba)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

wb.save('C:/Users/Arthur Coutinho/Desktop/final/final.xlsx')

print("Planilhas unidas com sucesso!")
